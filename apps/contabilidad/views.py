from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, View
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import Sum, Q, F, Prefetch
from django.utils.dateparse import parse_date
from apps.core.mixins import GestionMixin, AppPermisoMixin

class ContabilidadMixin(AppPermisoMixin):
    app_name = 'contabilidad'

from .models import PlanCuentas, AsientoContable, LineaAsiento, ConfiguracionContable, CentroCosto
from decimal import Decimal

# ---------------------------------------------------------------------------
# Plan de Cuentas
# ---------------------------------------------------------------------------

class PlanCuentasListView(ContabilidadMixin, ListView):
    model = PlanCuentas
    template_name = 'admin/contabilidad/plan_list.html'
    context_object_name = 'cuentas'
    paginate_by = 50

    def get_queryset(self):
        qs = PlanCuentas.objects.select_related('parent')
        q = self.request.GET.get('q')
        nivel = self.request.GET.get('nivel')
        if nivel:
            qs = qs.filter(nivel=nivel)
        if q:
            qs = qs.filter(codigo__icontains=q) | qs.filter(nombre__icontains=q)
        tipo = self.request.GET.get('tipo')
        if tipo:
            qs = qs.filter(tipo=tipo)
        return qs
        


    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['tipos'] = PlanCuentas.TIPO_CHOICES
        ctx['titulo'] = 'Plan de Cuentas'
        return ctx


class PlanCuentasCreateView(ContabilidadMixin, CreateView):
    model = PlanCuentas
    template_name = 'admin/contabilidad/plan_form.html'
    fields = ['codigo', 'nombre', 'tipo', 'nivel', 'parent', 'descripcion', 'acepta_movimientos', 'activa']
    success_url = reverse_lazy('contabilidad:plan_list')

    def form_valid(self, form):
        messages.success(self.request, 'Cuenta creada exitosamente.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Nueva Cuenta Contable'
        return ctx


class PlanCuentasUpdateView(ContabilidadMixin, UpdateView):
    model = PlanCuentas
    template_name = 'admin/contabilidad/plan_form.html'
    fields = ['codigo', 'nombre', 'tipo', 'nivel', 'parent', 'descripcion', 'acepta_movimientos', 'activa']
    success_url = reverse_lazy('contabilidad:plan_list')

    def form_valid(self, form):
        messages.success(self.request, 'Cuenta actualizada exitosamente.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f'Editar Cuenta: {self.object}'
        return ctx


class PlanCuentasDeleteView(ContabilidadMixin, DeleteView):
    model = PlanCuentas
    template_name = 'admin/confirm_delete.html'
    success_url = reverse_lazy('contabilidad:plan_list')

    def form_valid(self, form):
        messages.success(self.request, 'Cuenta eliminada.')
        return super().form_valid(form)



# ---------------------------------------------------------------------------
# Configuración Contable (singleton)
# ---------------------------------------------------------------------------



class ConfiguracionContableView(ContabilidadMixin, View):
    template_name = 'admin/contabilidad/configuracion_form.html'

    def _get_form_class(self):
        from django import forms

        class ConfigForm(forms.ModelForm):
            class Meta:
                model = ConfiguracionContable
                fields = [
                    'cuenta_cxc', 'cuenta_cxp',
                    'cuenta_documentos_por_pagar',
                    'cuenta_iva_debito', 'cuenta_iva_credito',
                    'cuenta_ventas_default', 'cuenta_compras_default',
                    'cuenta_sueldos_operacional', 'cuenta_sueldos_administrativo',
                    'cuenta_impuestos_sii',
                    'cuenta_afp_por_pagar', 'cuenta_salud_por_pagar',
                    'cuenta_sueldos_por_pagar',
                    'cuenta_anticipos_trabajadores', 'cuenta_anticipos_proveedores',
                    'cuenta_patrimonio_apertura',
                    'cuenta_honorarios_default', 'cuenta_retenciones_honorarios',
                ]
            # Editamos el init para limitar las opciones de cuentas a solo aquellas que son de nivel 4 y si son de gastos o ingresos según corresponda
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                cuentas = PlanCuentas.objects.filter(activa=True, nivel=4).order_by('codigo')
                self.fields['cuenta_cxc'].queryset = cuentas.filter(tipo='activo')
                self.fields['cuenta_cxp'].queryset = cuentas.filter(tipo='pasivo')
                self.fields['cuenta_documentos_por_pagar'].queryset = cuentas.filter(tipo='pasivo')
                self.fields['cuenta_iva_debito'].queryset = cuentas.filter(tipo='pasivo')
                self.fields['cuenta_iva_credito'].queryset = cuentas.filter(tipo='activo')
                self.fields['cuenta_ventas_default'].queryset = cuentas.filter(tipo='ingreso')
                self.fields['cuenta_compras_default'].queryset = cuentas.filter(tipo='gasto')
                self.fields['cuenta_sueldos_operacional'].queryset = cuentas.filter(tipo='costo')
                self.fields['cuenta_sueldos_administrativo'].queryset = cuentas.filter(tipo='gasto')
                self.fields['cuenta_impuestos_sii'].queryset = cuentas.filter(tipo='pasivo')
                self.fields['cuenta_afp_por_pagar'].queryset = cuentas.filter(tipo='pasivo')
                self.fields['cuenta_salud_por_pagar'].queryset = cuentas.filter(tipo='pasivo')
                self.fields['cuenta_sueldos_por_pagar'].queryset = cuentas.filter(tipo='pasivo')
                self.fields['cuenta_anticipos_trabajadores'].queryset = cuentas.filter(tipo='activo')
                self.fields['cuenta_anticipos_proveedores'].queryset = cuentas.filter(tipo='activo')
                self.fields['cuenta_patrimonio_apertura'].queryset = cuentas.filter(tipo='patrimonio')
                self.fields['cuenta_honorarios_default'].queryset = cuentas.filter(tipo__in=['gasto', 'costo'])
                self.fields['cuenta_retenciones_honorarios'].queryset = cuentas.filter(tipo='pasivo')

        return ConfigForm

    def get(self, request):
        from django.shortcuts import render
        obj = ConfiguracionContable.get()
        form = self._get_form_class()(instance=obj)
        return render(request, self.template_name, {'form': form, 'titulo': 'Configuración Contable'})

    def post(self, request):
        from django.shortcuts import render
        obj = ConfiguracionContable.get()
        form = self._get_form_class()(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuración guardada.')
            return redirect('contabilidad:configuracion')
        return render(request, self.template_name, {'form': form, 'titulo': 'Configuración Contable'})


# ---------------------------------------------------------------------------
# Libro Diario — CRUD Asientos
# ---------------------------------------------------------------------------

class LibroDiarioListView(ContabilidadMixin, ListView):
    model = AsientoContable
    template_name = 'admin/contabilidad/diario_list.html'
    context_object_name = 'asientos'
    paginate_by = 40

    def get_queryset(self):
        qs = AsientoContable.objects.prefetch_related('lineas')
        tipo = self.request.GET.get('tipo')
        estado = self.request.GET.get('estado')
        desde = self.request.GET.get('desde')
        hasta = self.request.GET.get('hasta')
        if tipo:
            qs = qs.filter(tipo=tipo)
        if estado:
            qs = qs.filter(estado=estado)
        if desde:
            qs = qs.filter(fecha__gte=parse_date(desde))
        if hasta:
            qs = qs.filter(fecha__lte=parse_date(hasta))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Libro Diario'
        ctx['tipo_choices'] = AsientoContable.TIPO_CHOICES
        ctx['estado_choices'] = AsientoContable.ESTADO_CHOICES
        return ctx


class AsientoDetailView(ContabilidadMixin, DetailView):
    model = AsientoContable
    template_name = 'admin/contabilidad/asiento_detail.html'
    context_object_name = 'asiento'

    def get_queryset(self):
        return super().get_queryset().prefetch_related(
            Prefetch(
                'lineas',
                queryset=LineaAsiento.objects.select_related(
                    'cuenta',
                    'centro_costo',
                ),
            )
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f'Asiento {self.object.numero}'
        return ctx


class AsientoCreateView(ContabilidadMixin, View):
    template_name = 'admin/contabilidad/asiento_form.html'

    def _forms(self, data=None):
        from django import forms
        from django.forms import inlineformset_factory

        class AsientoForm(forms.ModelForm):
            class Meta:
                model = AsientoContable
                fields = ['fecha', 'descripcion', 'tipo']
                widgets = {
                    'fecha': forms.DateInput(attrs={'type': 'date', 'class': 'form-control form-control-sm'}),
                    'descripcion': forms.TextInput(attrs={'class': 'form-control form-control-sm'}),
                    'tipo': forms.Select(attrs={'class': 'form-select form-select-sm'}),
                }

        class LineaForm(forms.ModelForm):
            class Meta:
                model = LineaAsiento
                fields = ['cuenta', 'descripcion', 'debe', 'haber', 'centro_costo']
                widgets = {
                    'cuenta': forms.Select(attrs={'class': 'form-select form-select-sm cuenta-select'}),
                    'descripcion': forms.TextInput(attrs={'class': 'form-control form-control-sm'}),
                    'debe': forms.NumberInput(attrs={'class': 'form-control form-control-sm text-end', 'step': '0.01', 'min': '0'}),
                    'haber': forms.NumberInput(attrs={'class': 'form-control form-control-sm text-end', 'step': '0.01', 'min': '0'}),
                    'centro_costo': forms.Select(attrs={'class': 'centro-select'}),
                }

        LineaFormSet = inlineformset_factory(AsientoContable, LineaAsiento, form=LineaForm, extra=0, can_delete=True, min_num=2)
        form = AsientoForm(data)
        formset = LineaFormSet(data)
        return form, formset

    def get(self, request):
        from django.shortcuts import render
        form, formset = self._forms()
        cuentas = PlanCuentas.objects.filter(activa=True, acepta_movimientos=True).order_by('codigo')
        centros_costo = CentroCosto.objects.filter(activo=True).order_by('codigo')
        return render(request, self.template_name, {
            'form': form, 'formset': formset, 'cuentas': cuentas,
            'centros_costo': centros_costo,
            'titulo': 'Nuevo Asiento Manual',
        })

    def post(self, request):
        from django.shortcuts import render
        from django.forms import inlineformset_factory

        class AsientoForm(__import__('django').forms.ModelForm):
            class Meta:
                model = AsientoContable
                fields = ['fecha', 'descripcion', 'tipo']

        class LineaForm(__import__('django').forms.ModelForm):
            class Meta:
                model = LineaAsiento
                fields = ['cuenta', 'descripcion', 'debe', 'haber', 'centro_costo']

        LineaFormSet = inlineformset_factory(AsientoContable, LineaAsiento, form=LineaForm, extra=0, can_delete=True)
        form = AsientoForm(request.POST)
        formset = LineaFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            # Validate balance before persisting
            total_debe = Decimal('0')
            total_haber = Decimal('0')
            for f in formset.forms:
                if f.cleaned_data and not f.cleaned_data.get('DELETE', False):
                    total_debe += f.cleaned_data.get('debe', Decimal('0')) or Decimal('0')
                    total_haber += f.cleaned_data.get('haber', Decimal('0')) or Decimal('0')
            if total_debe != total_haber:
                messages.error(
                    request,
                    f'El asiento no está cuadrado (Debe={total_debe} / Haber={total_haber}). '
                    'Corrija los montos antes de guardar.'
                )
            else:
                asiento = form.save(commit=False)
                asiento.creado_por = request.user
                asiento.save()
                bound_formset = LineaFormSet(request.POST, instance=asiento)
                if bound_formset.is_valid():
                    bound_formset.save()
                    messages.success(request, f'Asiento {asiento.numero} creado.')
                    return redirect('contabilidad:asiento_detail', pk=asiento.pk)
                else:
                    asiento.delete()
                    formset = bound_formset

        cuentas = PlanCuentas.objects.filter(activa=True, acepta_movimientos=True).order_by('codigo')
        centros_costo = CentroCosto.objects.filter(activo=True).order_by('codigo')
        return render(request, self.template_name, {
            'form': form, 'formset': formset, 'cuentas': cuentas,
            'centros_costo': centros_costo,
            'titulo': 'Nuevo Asiento Manual',
        })


class AsientoUpdateView(ContabilidadMixin, View):
    template_name = 'admin/contabilidad/asiento_form.html'

    def _get_asiento(self, pk):
        return get_object_or_404(AsientoContable, pk=pk, estado='borrador')

    def _formset_class(self):
        from django import forms
        from django.forms import inlineformset_factory

        class LineaForm(forms.ModelForm):
            class Meta:
                model = LineaAsiento
                fields = ['cuenta', 'descripcion', 'debe', 'haber', 'centro_costo']
                widgets = {
                    'cuenta': forms.Select(attrs={'class': 'form-select form-select-sm cuenta-select'}),
                    'descripcion': forms.TextInput(attrs={'class': 'form-control form-control-sm'}),
                    'debe': forms.NumberInput(attrs={'class': 'form-control form-control-sm text-end', 'step': '0.01', 'min': '0'}),
                    'haber': forms.NumberInput(attrs={'class': 'form-control form-control-sm text-end', 'step': '0.01', 'min': '0'}),
                    'centro_costo': forms.Select(attrs={'class': 'centro-select'}),
                }

        return inlineformset_factory(AsientoContable, LineaAsiento, form=LineaForm, extra=1, can_delete=True)

    def get(self, request, pk):
        from django import forms
        from django.shortcuts import render
        asiento = self._get_asiento(pk)

        class AsientoForm(forms.ModelForm):
            class Meta:
                model = AsientoContable
                fields = ['fecha', 'descripcion', 'tipo']
                widgets = {
                    'fecha': forms.DateInput(attrs={'type': 'date', 'class': 'form-control form-control-sm'}),
                    'descripcion': forms.TextInput(attrs={'class': 'form-control form-control-sm'}),
                    'tipo': forms.Select(attrs={'class': 'form-select form-select-sm'}),
                }

        form = AsientoForm(instance=asiento)
        formset = self._formset_class()(instance=asiento)
        cuentas = PlanCuentas.objects.filter(activa=True, acepta_movimientos=True).order_by('codigo')
        centros_costo = CentroCosto.objects.filter(activo=True).order_by('codigo')
        return render(request, self.template_name, {
            'form': form, 'formset': formset, 'cuentas': cuentas,
            'centros_costo': centros_costo,
            'asiento': asiento, 'titulo': f'Editar Asiento {asiento.numero}',
        })

    def post(self, request, pk):
        from django import forms
        from django.shortcuts import render
        asiento = self._get_asiento(pk)

        class AsientoForm(forms.ModelForm):
            class Meta:
                model = AsientoContable
                fields = ['fecha', 'descripcion', 'tipo']

        form = AsientoForm(request.POST, instance=asiento)
        formset = self._formset_class()(request.POST, instance=asiento)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Asiento actualizado.')
            return redirect('contabilidad:asiento_detail', pk=asiento.pk)

        cuentas = PlanCuentas.objects.filter(activa=True, acepta_movimientos=True).order_by('codigo')
        centros_costo = CentroCosto.objects.filter(activo=True).order_by('codigo')
        return render(request, self.template_name, {
            'form': form, 'formset': formset, 'cuentas': cuentas,
            'centros_costo': centros_costo,
            'asiento': asiento, 'titulo': f'Editar Asiento {asiento.numero}',
        })


class AsientoConfirmarView(ContabilidadMixin, View):
    def post(self, request, pk):
        asiento = get_object_or_404(AsientoContable, pk=pk)
        if asiento.estado != 'borrador':
            messages.warning(request, 'Solo los asientos en borrador pueden confirmarse.')
        elif not asiento.esta_cuadrado:
            messages.error(request, f'El asiento no está cuadrado (Debe={asiento.total_debe} / Haber={asiento.total_haber}).')
        elif asiento.lineas.count() == 0:
            messages.error(request, 'El asiento no tiene líneas.')
        else:
            asiento.estado = 'confirmado'
            asiento.save(update_fields=['estado'])
            messages.success(request, f'Asiento {asiento.numero} confirmado.')
        return redirect('contabilidad:asiento_detail', pk=asiento.pk)


class AsientoAnularView(ContabilidadMixin, View):
    def post(self, request, pk):
        asiento = get_object_or_404(AsientoContable, pk=pk)
        if asiento.estado == 'anulado':
            messages.warning(request, 'El asiento ya está anulado.')
        else:
            asiento.estado = 'anulado'
            asiento.save(update_fields=['estado'])
            messages.success(request, f'Asiento {asiento.numero} anulado.')
        return redirect('contabilidad:asiento_detail', pk=asiento.pk)

class AsientoDeleteView(ContabilidadMixin, DeleteView):
    model = AsientoContable
    template_name = 'admin/confirm_delete.html'
    success_url = reverse_lazy('contabilidad:diario_list')

    def form_valid(self, form):
        if self.object.estado == 'confirmado':
            messages.warning(self.request, 'Los asientos confirmados no pueden eliminarse. Anúlelo primero.')
            return redirect('contabilidad:asiento_detail', pk=self.object.pk)
        messages.success(self.request, f'Asiento {self.object.numero} eliminado.')
        return super().form_valid(form)


class AsientosExcelView(ContabilidadMixin, View):
    """
    Descarga el libro diario con sus líneas como archivo Excel (.xlsx).
    Acepta los mismos filtros GET que LibroDiarioListView: tipo, estado, desde, hasta.
    Genera dos hojas:
        - Resumen: una fila por asiento
        - Detalle Líneas: una fila por línea de asiento
    """

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.utils import timezone as tz

        # ── Filtros (idénticos a LibroDiarioListView) ─────────────────────
        qs = (
            AsientoContable.objects
            .prefetch_related('lineas__cuenta', 'lineas__centro_costo')
            .order_by('fecha', 'numero')
        )
        tipo   = request.GET.get('tipo')
        estado = request.GET.get('estado')
        desde  = request.GET.get('desde')
        hasta  = request.GET.get('hasta')
        if tipo:
            qs = qs.filter(tipo=tipo)
        if estado:
            qs = qs.filter(estado=estado)
        if desde:
            qs = qs.filter(fecha__gte=parse_date(desde))
        if hasta:
            qs = qs.filter(fecha__lte=parse_date(hasta))

        # ── Estilos ───────────────────────────────────────────────────────
        font_header  = Font(bold=True, color='FFFFFF')
        font_bold    = Font(bold=True)
        fill_header  = PatternFill('solid', fgColor='1F3864')
        fill_asiento = PatternFill('solid', fgColor='D9E1F2')
        fill_anulado = PatternFill('solid', fgColor='FFD7D7')
        fill_total   = PatternFill('solid', fgColor='EBF3E8')
        align_center = Alignment(horizontal='center', vertical='center')
        align_right  = Alignment(horizontal='right')
        fmt_num      = '#,##0.00'
        fmt_date     = 'DD/MM/YYYY'

        wb = openpyxl.Workbook()

        # ═══════════════════════════════════════════════════════════════════
        # Hoja 1 — Resumen (una fila por asiento)
        # ═══════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Resumen'

        h1 = ['N° Asiento', 'Fecha', 'Descripción', 'Tipo', 'Estado',
              'Total Debe', 'Total Haber', 'Cuadrado']
        for col, val in enumerate(h1, 1):
            c = ws1.cell(row=1, column=col, value=val)
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_center

        for col, w in enumerate([18, 14, 55, 28, 14, 18, 18, 12], 1):
            ws1.column_dimensions[get_column_letter(col)].width = w
        ws1.row_dimensions[1].height = 20

        for r, asiento in enumerate(qs, start=2):
            fill = fill_anulado if asiento.estado == 'anulado' else None
            cuadrado = asiento.esta_cuadrado
            fila = [
                asiento.numero,
                asiento.fecha,
                asiento.descripcion,
                asiento.get_tipo_display(),
                asiento.get_estado_display(),
                float(asiento.total_debe),
                float(asiento.total_haber),
                'Sí' if cuadrado else 'No',
            ]
            for col, val in enumerate(fila, 1):
                cell = ws1.cell(row=r, column=col, value=val)
                if fill:
                    cell.fill = fill
                if col == 2:
                    cell.number_format = fmt_date
                if col in (6, 7):
                    cell.number_format = fmt_num
                    cell.alignment = align_right
                if col == 8:
                    cell.alignment = align_center
                    cell.font = Font(bold=True, color='006100' if cuadrado else 'C00000')

        ws1.freeze_panes = 'A2'
        ws1.auto_filter.ref = f'A1:{get_column_letter(len(h1))}1'

        # ═══════════════════════════════════════════════════════════════════
        # Hoja 2 — Detalle líneas (una fila por línea de asiento)
        # ═══════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Detalle Líneas')

        h2 = ['N° Asiento', 'Fecha', 'Desc. Asiento', 'Tipo', 'Estado',
              'Orden', 'Cód. Cuenta', 'Nombre Cuenta', 'Desc. Línea', 'Centro Costo', 'Debe', 'Haber']
        for col, val in enumerate(h2, 1):
            c = ws2.cell(row=1, column=col, value=val)
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_center

        for col, w in enumerate([18, 14, 50, 28, 14, 8, 15, 40, 50, 20, 18, 18], 1):
            ws2.column_dimensions[get_column_letter(col)].width = w
        ws2.row_dimensions[1].height = 20

        row = 2
        for asiento in qs:
            lineas = list(asiento.lineas.all())
            fill = fill_anulado if asiento.estado == 'anulado' else None
            for linea in lineas:
                fila = [
                    asiento.numero,
                    asiento.fecha,
                    asiento.descripcion,
                    asiento.get_tipo_display(),
                    asiento.get_estado_display(),
                    linea.orden,
                    linea.cuenta.codigo,
                    linea.cuenta.nombre,
                    linea.descripcion,
                    str(linea.centro_costo) if linea.centro_costo_id else '',
                    float(linea.debe),
                    float(linea.haber),
                ]
                for col, val in enumerate(fila, 1):
                    cell = ws2.cell(row=row, column=col, value=val)
                    if fill:
                        cell.fill = fill
                    if col == 2:
                        cell.number_format = fmt_date
                    if col in (11, 12):
                        cell.number_format = fmt_num
                        cell.alignment = align_right
                row += 1

            # Fila subtotal por asiento
            if lineas:
                for col in range(1, len(h2) + 1):
                    ws2.cell(row=row, column=col).fill = fill_total
                lbl = ws2.cell(row=row, column=9, value=f'Subtotal {asiento.numero}')
                lbl.font = font_bold
                lbl.fill = fill_total
                for col, val in [(11, float(asiento.total_debe)), (12, float(asiento.total_haber))]:
                    c = ws2.cell(row=row, column=col, value=val)
                    c.font = font_bold
                    c.fill = fill_total
                    c.number_format = fmt_num
                    c.alignment = align_right
                row += 1

        ws2.freeze_panes = 'A2'
        ws2.auto_filter.ref = f'A1:{get_column_letter(len(h2))}1'

        # ── Respuesta ─────────────────────────────────────────────────────
        filename = f'libro_diario_{tz.now().strftime("%Y%m%d")}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

# ---------------------------------------------------------------------------
# Reportes Contables
# ---------------------------------------------------------------------------

class LibroMayorView(ContabilidadMixin, View):
    template_name = 'admin/contabilidad/libro_mayor.html'

    def get(self, request):
        from django.shortcuts import render
        desde = request.GET.get('desde')
        hasta = request.GET.get('hasta')
        cuenta_id = request.GET.get('cuenta')

        cuentas = PlanCuentas.objects.filter(activa=True, acepta_movimientos=True).order_by('codigo')
        movimientos = []
        cuenta_sel = None

        if cuenta_id:
            cuenta_sel = get_object_or_404(PlanCuentas, pk=cuenta_id)
            qs = LineaAsiento.objects.filter(
                asiento__estado='confirmado',
                cuenta=cuenta_sel,
            ).select_related('asiento').order_by('asiento__fecha', 'asiento__numero', 'orden')
            if desde:
                qs = qs.filter(asiento__fecha__gte=parse_date(desde))
            if hasta:
                qs = qs.filter(asiento__fecha__lte=parse_date(hasta))

            saldo = 0
            for linea in qs:
                saldo += float(linea.debe) - float(linea.haber)
                movimientos.append({
                    'linea': linea,
                    'saldo': saldo,
                })

        return render(request, self.template_name, {
            'titulo': 'Libro Mayor',
            'cuentas': cuentas,
            'cuenta_sel': cuenta_sel,
            'movimientos': movimientos,
        })


class BalanceComprobacionView(ContabilidadMixin, View):
    template_name = 'admin/contabilidad/balance_comprobacion.html'

    def get(self, request):
        from django.shortcuts import render
        desde = request.GET.get('desde')
        hasta = request.GET.get('hasta')

        qs = LineaAsiento.objects.filter(asiento__estado='confirmado').select_related('cuenta')
        if desde:
            qs = qs.filter(asiento__fecha__gte=parse_date(desde))
        if hasta:
            qs = qs.filter(asiento__fecha__lte=parse_date(hasta))

        from collections import defaultdict
        from decimal import Decimal
        data = defaultdict(lambda: {'debe': Decimal('0'), 'haber': Decimal('0'), 'cuenta': None})
        for linea in qs:
            key = linea.cuenta_id
            data[key]['debe'] += linea.debe
            data[key]['haber'] += linea.haber
            data[key]['cuenta'] = linea.cuenta

        filas = []
        total_debe = Decimal('0')
        total_haber = Decimal('0')
        for key in sorted(data.keys(), key=lambda k: data[k]['cuenta'].codigo if data[k]['cuenta'] else ''):
            row = data[key]
            saldo = row['debe'] - row['haber']
            filas.append({
                'cuenta': row['cuenta'],
                'debe': row['debe'],
                'haber': row['haber'],
                'saldo': saldo,
            })
            total_debe += row['debe']
            total_haber += row['haber']

        return render(request, self.template_name, {
            'titulo': 'Balance de Comprobación',
            'filas': filas,
            'total_debe': total_debe,
            'total_haber': total_haber,
        })


class BalanceGeneralView(ContabilidadMixin, View):
    template_name = 'admin/contabilidad/balance_general.html'

    def get(self, request):
        from django.shortcuts import render
        hasta = request.GET.get('hasta')

        qs = LineaAsiento.objects.filter(asiento__estado='confirmado').select_related('cuenta')
        if hasta:
            qs = qs.filter(asiento__fecha__lte=parse_date(hasta))

        from collections import defaultdict
        from decimal import Decimal
        saldos = defaultdict(Decimal)
        cuentas_map = {}
        for linea in qs:
            saldos[linea.cuenta_id] += linea.debe - linea.haber
            cuentas_map[linea.cuenta_id] = linea.cuenta

        activos, pasivos, patrimonio = [], [], []
        total_activo = Decimal('0')
        total_pasivo = Decimal('0')
        total_patrimonio = Decimal('0')
        resultado_ejercicio = Decimal('0')

        for cid, saldo in saldos.items():
            if saldo == 0:
                continue
            cuenta = cuentas_map[cid]
            if cuenta.tipo == 'activo':
                row = {'cuenta': cuenta, 'saldo': saldo}
                activos.append(row)
                total_activo += saldo
            elif cuenta.tipo == 'pasivo':
                monto = -saldo
                row = {'cuenta': cuenta, 'saldo': monto}
                pasivos.append(row)
                total_pasivo += monto
            elif cuenta.tipo == 'patrimonio':
                monto = -saldo
                row = {'cuenta': cuenta, 'saldo': monto}
                patrimonio.append(row)
                total_patrimonio += monto
            elif cuenta.tipo in ('ingreso', 'costo', 'gasto', 'socio'):
                resultado_ejercicio -= saldo

        activos.sort(key=lambda r: r['cuenta'].codigo)
        pasivos.sort(key=lambda r: r['cuenta'].codigo)
        patrimonio.sort(key=lambda r: r['cuenta'].codigo)

        if resultado_ejercicio:
            patrimonio.append({
                'cuenta': None,
                'saldo': resultado_ejercicio,
                'es_resultado_ejercicio': True,
            })
            total_patrimonio += resultado_ejercicio

        return render(request, self.template_name, {
            'titulo': 'Balance General',
            'activos': activos, 'total_activo': total_activo,
            'pasivos': pasivos, 'total_pasivo': total_pasivo,
            'patrimonio': patrimonio, 'total_patrimonio': total_patrimonio,
            'resultado_ejercicio': resultado_ejercicio,
            'total_pasivo_patrimonio': total_pasivo + total_patrimonio,
        })


class EstadoResultadosView(ContabilidadMixin, View):
    template_name = 'admin/contabilidad/estado_resultados.html'

    def _ruta_cuenta(self, cuenta):
        return cuenta.get_ruta().lower()

    def _es_otros_ingresos(self, cuenta):
        return cuenta.tipo == 'ingreso' and 'otros ingresos' in self._ruta_cuenta(cuenta)

    def _es_impuesto_renta(self, cuenta):
        ruta = self._ruta_cuenta(cuenta)
        return (
            cuenta.tipo in ('gasto', 'costo') and
            'impuesto' in ruta and
            ('renta' in ruta or 'primera categoria' in ruta or 'primera categoría' in ruta)
        )

    def _es_otro_gasto(self, cuenta):
        ruta = self._ruta_cuenta(cuenta)
        return (
            cuenta.tipo in ('gasto', 'costo') and
            not self._es_impuesto_renta(cuenta) and
            ('financiamiento' in ruta or 'intereses' in ruta)
        )

    def get(self, request):
        from django.shortcuts import render
        desde = request.GET.get('desde')
        hasta = request.GET.get('hasta')

        qs = LineaAsiento.objects.filter(asiento__estado='confirmado').select_related(
            'cuenta',
            'cuenta__parent',
            'cuenta__parent__parent',
            'cuenta__parent__parent__parent',
        )
        if desde:
            qs = qs.filter(asiento__fecha__gte=parse_date(desde))
        if hasta:
            qs = qs.filter(asiento__fecha__lte=parse_date(hasta))

        from collections import defaultdict
        from decimal import Decimal
        saldos = defaultdict(Decimal)
        cuentas_map = {}
        for linea in qs:
            saldos[linea.cuenta_id] += linea.haber - linea.debe  # ingresos son saldo haber
            cuentas_map[linea.cuenta_id] = linea.cuenta

        ingresos, costos, gastos, otros, impuestos_renta = [], [], [], [], []
        total_ingresos = Decimal('0')
        total_costos = Decimal('0')
        total_gastos = Decimal('0')
        total_otros = Decimal('0')
        total_impuesto_renta = Decimal('0')

        for cid, saldo in saldos.items():
            if saldo == 0:
                continue
            cuenta = cuentas_map[cid]
            if self._es_impuesto_renta(cuenta):
                monto = -saldo
                impuestos_renta.append({'cuenta': cuenta, 'saldo': monto})
                total_impuesto_renta += monto
            elif self._es_otros_ingresos(cuenta) or self._es_otro_gasto(cuenta):
                monto = saldo
                otros.append({'cuenta': cuenta, 'saldo': monto})
                total_otros += monto
            elif cuenta.tipo == 'ingreso':
                monto = saldo
                row = {'cuenta': cuenta, 'saldo': monto}
                ingresos.append(row)
                total_ingresos += monto
            elif cuenta.tipo == 'costo':
                monto = -saldo
                row = {'cuenta': cuenta, 'saldo': monto}
                costos.append(row)
                total_costos += monto
            elif cuenta.tipo in ('gasto', 'socio'):
                monto = -saldo
                row = {'cuenta': cuenta, 'saldo': monto}
                gastos.append(row)
                total_gastos += monto

        ingresos.sort(key=lambda r: r['cuenta'].codigo)
        costos.sort(key=lambda r: r['cuenta'].codigo)
        gastos.sort(key=lambda r: r['cuenta'].codigo)
        otros.sort(key=lambda r: r['cuenta'].codigo)
        impuestos_renta.sort(key=lambda r: r['cuenta'].codigo)

        utilidad_bruta = total_ingresos - total_costos
        resultado_operacional = utilidad_bruta - total_gastos
        resultado_antes_impuestos = resultado_operacional + total_otros
        utilidad_ejercicio = resultado_antes_impuestos - total_impuesto_renta

        return render(request, self.template_name, {
            'titulo': 'Estado de Resultados',
            'ingresos': ingresos, 'total_ingresos': total_ingresos,
            'costos': costos, 'total_costos': total_costos,
            'gastos': gastos, 'total_gastos': total_gastos,
            'otros': otros, 'total_otros': total_otros,
            'impuestos_renta': impuestos_renta, 'total_impuesto_renta': total_impuesto_renta,
            'utilidad_bruta': utilidad_bruta,
            'resultado_operacional': resultado_operacional,
            'resultado_antes_impuestos': resultado_antes_impuestos,
            'utilidad_ejercicio': utilidad_ejercicio,
        })



# ---------------------------------------------------------------------------
# Centros de Costo
# ---------------------------------------------------------------------------

class CentroCostoListView(ContabilidadMixin, ListView):
    model = CentroCosto
    template_name = 'admin/contabilidad/centro_costo_list.html'
    context_object_name = 'centros'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Centros de Costo'
        return ctx


class CentroCostoCreateView(ContabilidadMixin, CreateView):
    model = CentroCosto
    template_name = 'admin/contabilidad/centro_costo_form.html'
    fields = ['codigo', 'nombre', 'descripcion', 'presupuesto_mensual', 'activo']
    success_url = reverse_lazy('contabilidad:centro_list')

    def form_valid(self, form):
        messages.success(self.request, 'Centro de costo creado.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Nuevo Centro de Costo'
        return ctx


class CentroCostoUpdateView(ContabilidadMixin, UpdateView):
    model = CentroCosto
    template_name = 'admin/contabilidad/centro_costo_form.html'
    fields = ['codigo', 'nombre', 'descripcion', 'presupuesto_mensual', 'activo']
    success_url = reverse_lazy('contabilidad:centro_list')

    def form_valid(self, form):
        messages.success(self.request, 'Centro de costo actualizado.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f'Editar: {self.object}'
        return ctx


class CentroCostoDetalleView(ContabilidadMixin, View):
    """Reporte detallado de un centro de costo: ingresos, egresos y todos los movimientos."""
    template_name = 'admin/contabilidad/centro_costo_detalle.html'

    def get(self, request, pk):
        from django.shortcuts import render
        from decimal import Decimal
        from apps.clientes.models import DetalleFacturaEmitida
        from apps.proveedores.models import DetalleFacturaRecibida
        from apps.rendiciones.models import DetalleRendicion
        from apps.rrhh.models import Remuneracion
        from apps.proyectos.models import CostoProyecto

        centro = get_object_or_404(CentroCosto, pk=pk)
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')

        # ── Facturas emitidas (ingresos) ──────────────────────────────────
        fe_qs = (
            DetalleFacturaEmitida.objects
            .filter(centro_costo=centro)
            .exclude(factura__estado='anulada')
            .select_related('factura', 'factura__cliente', 'cuenta_contable')
            .order_by('factura__fecha_emision')
        )
        if fecha_desde:
            fe_qs = fe_qs.filter(factura__fecha_emision__gte=fecha_desde)
        if fecha_hasta:
            fe_qs = fe_qs.filter(factura__fecha_emision__lte=fecha_hasta)
        facturas_emitidas = [
            {'obj': d, 'subtotal': d.cantidad * d.precio_unitario}
            for d in fe_qs
        ]
        total_ingresos = sum(r['subtotal'] for r in facturas_emitidas)

        # ── Facturas recibidas (egresos) ──────────────────────────────────
        fr_qs = (
            DetalleFacturaRecibida.objects
            .filter(centro_costo=centro)
            .exclude(factura__estado='anulada')
            .select_related('factura', 'factura__proveedor', 'cuenta_contable')
            .order_by('factura__fecha_emision')
        )
        if fecha_desde:
            fr_qs = fr_qs.filter(factura__fecha_emision__gte=fecha_desde)
        if fecha_hasta:
            fr_qs = fr_qs.filter(factura__fecha_emision__lte=fecha_hasta)
        facturas_recibidas = [
            {'obj': d, 'subtotal': d.cantidad * d.precio_unitario}
            for d in fr_qs
        ]
        total_fact_recibidas = sum(r['subtotal'] for r in facturas_recibidas)

        # ── Rendiciones (egresos) ─────────────────────────────────────────
        rend_qs = (
            DetalleRendicion.objects
            .filter(centro_costo=centro, rendicion__estado='aprobado')
            .select_related('rendicion', 'rendicion__trabajador', 'cuenta_contable')
            .order_by('fecha_gasto')
        )
        if fecha_desde:
            rend_qs = rend_qs.filter(fecha_gasto__gte=fecha_desde)
        if fecha_hasta:
            rend_qs = rend_qs.filter(fecha_gasto__lte=fecha_hasta)
        rendiciones = list(rend_qs)
        total_rendiciones = sum(r.monto for r in rendiciones)

        # ── Remuneraciones (egresos) ──────────────────────────────────────
        rem_qs = (
            Remuneracion.objects
            .filter(trabajador__centro_costo=centro)
            .exclude(estado='borrador')
            .select_related('trabajador')
            .order_by('periodo_anio', 'periodo_mes')
        )
        if fecha_desde:
            from django.utils.dateparse import parse_date as _pd
            d = _pd(fecha_desde)
            if d:
                rem_qs = rem_qs.filter(
                    periodo_anio__gte=d.year
                ).exclude(
                    periodo_anio=d.year, periodo_mes__lt=d.month
                )
        if fecha_hasta:
            from django.utils.dateparse import parse_date as _pd2
            h = _pd2(fecha_hasta)
            if h:
                rem_qs = rem_qs.filter(
                    periodo_anio__lte=h.year
                ).exclude(
                    periodo_anio=h.year, periodo_mes__gt=h.month
                )
        remuneraciones = list(rem_qs)
        total_remuneraciones = sum(r.sueldo_bruto for r in remuneraciones)

        # ── Costos de proyecto (egresos) ──────────────────────────────────
        cp_qs = (
            CostoProyecto.objects
            .filter(centro_costo=centro)
            .select_related('proyecto', 'cuenta_contable', 'proveedor')
            .order_by('fecha')
        )
        if fecha_desde:
            cp_qs = cp_qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            cp_qs = cp_qs.filter(fecha__lte=fecha_hasta)
        costos_proyecto = list(cp_qs)
        total_costos_proyecto = sum(c.monto for c in costos_proyecto)

        # ── Totals ────────────────────────────────────────────────────────
        total_egresos = total_fact_recibidas + total_rendiciones + total_remuneraciones + total_costos_proyecto
        resultado = total_ingresos - total_egresos

        return render(request, self.template_name, {
            'titulo': f'Detalle Centro de Costo: {centro}',
            'centro': centro,
            'fecha_desde': fecha_desde or '',
            'fecha_hasta': fecha_hasta or '',
            # ingresos
            'facturas_emitidas': facturas_emitidas,
            'total_ingresos': total_ingresos,
            # egresos
            'facturas_recibidas': facturas_recibidas,
            'total_fact_recibidas': total_fact_recibidas,
            'rendiciones': rendiciones,
            'total_rendiciones': total_rendiciones,
            'remuneraciones': remuneraciones,
            'total_remuneraciones': total_remuneraciones,
            'costos_proyecto': costos_proyecto,
            'total_costos_proyecto': total_costos_proyecto,
            # summary
            'total_egresos': total_egresos,
            'resultado': resultado,
        })


class CentroCostoReporteView(ContabilidadMixin, View):
    template_name = 'admin/contabilidad/centro_costo_reporte.html'

    def get(self, request):
        from django.shortcuts import render
        from decimal import Decimal
        from django.utils import timezone
        from django.db.models import F

        hoy = timezone.now()
        mes = int(request.GET.get('mes', hoy.month))
        anio = int(request.GET.get('anio', hoy.year))

        MESES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                 7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

        centros = CentroCosto.objects.filter(activo=True)

        # ── Gastos: detalles de facturas recibidas ──
        from apps.proveedores.models import DetalleFacturaRecibida
        gastos_qs = (
            DetalleFacturaRecibida.objects
            .filter(centro_costo__isnull=False,
                    factura__fecha_emision__month=mes,
                    factura__fecha_emision__year=anio)
            .exclude(factura__estado='anulada')
            .annotate(subtotal=F('cantidad') * F('precio_unitario'))
            .values('centro_costo_id')
            .annotate(total=Sum('subtotal'))
        )
        gastos_map = {r['centro_costo_id']: r['total'] for r in gastos_qs}

        # ── Ingresos: detalles de facturas emitidas ──
        from apps.clientes.models import DetalleFacturaEmitida
        ingresos_qs = (
            DetalleFacturaEmitida.objects
            .filter(centro_costo__isnull=False,
                    factura__fecha_emision__month=mes,
                    factura__fecha_emision__year=anio)
            .exclude(factura__estado='anulada')
            .annotate(subtotal=F('cantidad') * F('precio_unitario'))
            .values('centro_costo_id')
            .annotate(total=Sum('subtotal'))
        )
        ingresos_map = {r['centro_costo_id']: r['total'] for r in ingresos_qs}

        # ── Remuneraciones ──
        from apps.rrhh.models import Remuneracion
        rem_qs = (
            Remuneracion.objects
            .filter(trabajador__centro_costo__isnull=False,
                    periodo_mes=mes,
                    periodo_anio=anio)
            .values('trabajador__centro_costo_id')
            .annotate(total=Sum('sueldo_bruto'))
        )
        rem_map = {r['trabajador__centro_costo_id']: r['total'] for r in rem_qs}

        # ── Costos de proyecto ──
        from apps.proyectos.models import CostoProyecto
        costos_qs = (
            CostoProyecto.objects
            .filter(centro_costo__isnull=False,
                    fecha__month=mes,
                    fecha__year=anio)
            .values('centro_costo_id')
            .annotate(total=Sum('monto'))
        )
        costos_map = {r['centro_costo_id']: r['total'] for r in costos_qs}

        filas = []
        total_presupuesto = Decimal('0')
        total_gastos_fact = Decimal('0')
        total_ingresos = Decimal('0')
        total_remuneraciones = Decimal('0')
        total_costos_proy = Decimal('0')

        for c in centros:
            g  = gastos_map.get(c.pk, Decimal('0')) or Decimal('0')
            ig = ingresos_map.get(c.pk, Decimal('0')) or Decimal('0')
            r  = rem_map.get(c.pk, Decimal('0')) or Decimal('0')
            p  = costos_map.get(c.pk, Decimal('0')) or Decimal('0')
            egresos = g + r + p
            resultado = ig - egresos
            presup = c.presupuesto_mensual or Decimal('0')
            desv = None
            if presup > 0:
                desv = ((egresos - presup) / presup * 100).quantize(Decimal('0.1'))
            filas.append({
                'centro': c,
                'gastos_facturas': g,
                'ingresos': ig,
                'remuneraciones': r,
                'costos_proyecto': p,
                'egresos': egresos,
                'resultado': resultado,
                'presupuesto': presup,
                'desviacion': desv,
            })
            total_presupuesto += presup
            total_gastos_fact += g
            total_ingresos += ig
            total_remuneraciones += r
            total_costos_proy += p

        total_egresos = total_gastos_fact + total_remuneraciones + total_costos_proy
        total_resultado = total_ingresos - total_egresos
        desv_total = None
        if total_presupuesto > 0:
            desv_total = ((total_egresos - total_presupuesto) / total_presupuesto * 100).quantize(Decimal('0.1'))

        anios_disponibles = list(range(hoy.year - 2, hoy.year + 2))

        return render(request, self.template_name, {
            'titulo': f'Reporte Centros de Costo — {MESES[mes]} {anio}',
            'filas': filas,
            'mes': mes,
            'anio': anio,
            'mes_nombre': MESES[mes],
            'meses': MESES,
            'anios_disponibles': anios_disponibles,
            'total_presupuesto': total_presupuesto,
            'total_gastos_fact': total_gastos_fact,
            'total_ingresos': total_ingresos,
            'total_remuneraciones': total_remuneraciones,
            'total_costos_proy': total_costos_proy,
            'total_egresos': total_egresos,
            'total_resultado': total_resultado,
            'desv_total': desv_total,
        })

    
class InformeCentroCostoView(ContabilidadMixin, View):
    template_name = "admin/contabilidad/informe_centro_costo.html"

    def get(self, request, *args, **kwargs):
        centro_id = request.GET.get("centro_costo")
        fecha_desde = request.GET.get("fecha_desde")
        fecha_hasta = request.GET.get("fecha_hasta")

        centros = CentroCosto.objects.filter(activo=True).order_by("codigo")

        movimientos = LineaAsiento.objects.filter(
            centro_costo__isnull=False,
            asiento__estado="confirmado",
            cuenta__tipo__in=("ingreso", "costo", "gasto", "socio"),
        )

        if centro_id:
            movimientos = movimientos.filter(centro_costo_id=centro_id)

        if fecha_desde:
            movimientos = movimientos.filter(asiento__fecha__gte=fecha_desde)

        if fecha_hasta:
            movimientos = movimientos.filter(asiento__fecha__lte=fecha_hasta)

        resumen = (
            movimientos
            .values(
                "centro_costo_id",
                "centro_costo__codigo",
                "centro_costo__nombre",
                "centro_costo__presupuesto_mensual",
                "cuenta__tipo",
            )
            .annotate(
                total_debe=Sum("debe"),
                total_haber=Sum("haber"),
            )
            .order_by("centro_costo__codigo")
        )

        data = {}

        for item in resumen:
            debe = item["total_debe"] or Decimal("0.00")
            haber = item["total_haber"] or Decimal("0.00")
            centro_id_item = item["centro_costo_id"]
            fila = data.setdefault(centro_id_item, {
                "centro_costo_id": centro_id_item,
                "codigo": item["centro_costo__codigo"],
                "nombre": item["centro_costo__nombre"],
                "presupuesto": (
                    item["centro_costo__presupuesto_mensual"] or Decimal("0.00")
                ),
                "ingresos": Decimal("0.00"),
                "egresos": Decimal("0.00"),
            })

            if item["cuenta__tipo"] == "ingreso":
                fila["ingresos"] += haber - debe
            else:
                fila["egresos"] += debe - haber

        informe = []
        total_ingresos = Decimal("0.00")
        total_egresos = Decimal("0.00")

        for fila in data.values():
            fila["resultado"] = fila["ingresos"] - fila["egresos"]
            fila["diferencia"] = fila["presupuesto"] - fila["egresos"]
            informe.append(fila)
            total_ingresos += fila["ingresos"]
            total_egresos += fila["egresos"]

        informe.sort(key=lambda item: item["codigo"])
        total_resultado = total_ingresos - total_egresos

        context = {
            "centros": centros,
            "informe": informe,
            "total_ingresos": total_ingresos,
            "total_egresos": total_egresos,
            "total_resultado": total_resultado,
            "centro_id": centro_id,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
        }

        return render(request, self.template_name, context)


# ---------------------------------------------------------------------------
# Asiento de Apertura — Saldos Iniciales
# ---------------------------------------------------------------------------

class AperturaContableView(ContabilidadMixin, View):
    """
    Permite ingresar los saldos de apertura contable de la empresa.

    Crea (o reemplaza) un único AsientoContable de tipo 'apertura'. Al confirmarlo,
    todos los reportes (Balance General, Comprobación, Libro Mayor) lo incluyen
    automáticamente porque filtran por asiento__estado='confirmado', sin requerir
    ningún cambio en las vistas de reportes.

    Lógica de cuadre:
        DEBE  = activos con saldo
        HABER = pasivos con saldo
        La diferencia (DEBE - HABER) se registra automáticamente en
        ConfiguracionContable.cuenta_patrimonio_apertura (Resultados Acumulados / Capital).
    """
    template_name = 'admin/contabilidad/apertura.html'

    def _get_apertura(self):
        return AsientoContable.objects.filter(tipo='apertura').order_by('fecha').first()

    def get(self, request):
        from django.shortcuts import render
        import json
        apertura = self._get_apertura()
        cuentas = PlanCuentas.objects.filter(activa=True, acepta_movimientos=True).order_by('codigo')
        config = ConfiguracionContable.get()

        # Collect bank accounts linked to a PlanCuentas account (single source of truth: saldo_inicial)
        from apps.tesoreria.models import CuentaBancaria
        cuentas_bancarias = (
            CuentaBancaria.objects
            .filter(activa=True, cuenta_contable__isnull=False)
            .select_related('cuenta_contable', 'banco')
        )
        banco_cuenta_ids = {cb.cuenta_contable_id: cb for cb in cuentas_bancarias}

        # Pre-fill manual lines from existing apertura (excluding patrimonio and bank lines)
        patrimonio_id = config.cuenta_patrimonio_apertura_id if config else None
        lineas_existentes = []
        if apertura:
            for linea in apertura.lineas.select_related('cuenta').order_by('orden'):
                if linea.cuenta_id == patrimonio_id:
                    continue  # auto-balance line, recalculated on save
                if linea.cuenta_id in banco_cuenta_ids:
                    continue  # bank line, driven by saldo_inicial — don't show as manual
                lineas_existentes.append({
                    'cuenta_id': linea.cuenta_id,
                    'cuenta_str': str(linea.cuenta),
                    'monto': float(linea.debe if linea.debe > 0 else linea.haber),
                    'tipo': 'debe' if linea.debe > 0 else 'haber',
                    'from_banco': False,
                })

        cuentas_json = json.dumps([
            {'id': c.pk, 'label': str(c), 'tipo': c.tipo}
            for c in cuentas
            if c.pk not in banco_cuenta_ids  # bank accounts are shown separately, not in dropdown
        ])

        # Bank accounts summary for display (always shown, read-only)
        bancos_apertura = [
            {
                'cuenta_id': cb.cuenta_contable_id,
                'label': f'{cb.banco} — {cb.numero}',
                'cuenta_str': str(cb.cuenta_contable),
                'saldo_inicial': float(cb.saldo_inicial),
            }
            for cb in cuentas_bancarias
            if cb.saldo_inicial != 0
        ]

        return render(request, self.template_name, {
            'titulo': 'Saldos de Apertura',
            'apertura': apertura,
            'config': config,
            'cuentas_json': cuentas_json,
            'lineas_json': json.dumps(lineas_existentes),
            'bancos_apertura': bancos_apertura,
            'bancos_apertura_json': json.dumps(bancos_apertura),
            'fecha_apertura': apertura.fecha.isoformat() if apertura else '',
        })

    def post(self, request):
        from django.shortcuts import render, redirect
        from decimal import Decimal, InvalidOperation
        import json

        # ── Confirm existing borrador ──────────────────────────────────────
        if request.POST.get('action') == 'confirmar':
            apertura = self._get_apertura()
            if apertura and apertura.estado == 'borrador':
                if not apertura.esta_cuadrado:
                    messages.error(request, 'El asiento no está cuadrado (Debe ≠ Haber). Corrija los montos.')
                    return redirect('contabilidad:apertura')
                apertura.estado = 'confirmado'
                apertura.save(update_fields=['estado'])
                messages.success(request, f'Asiento de apertura {apertura.numero} confirmado. Todos los reportes lo incluirán.')
            return redirect('contabilidad:apertura')

        # ── Save / replace borrador ────────────────────────────────────────
        config = ConfiguracionContable.get()

        fecha_str = request.POST.get('fecha', '').strip()
        fecha = parse_date(fecha_str)
        if not fecha:
            messages.error(request, 'Ingrese una fecha de apertura válida.')
            return redirect('contabilidad:apertura')

        cuenta_ids = request.POST.getlist('cuenta_id')
        montos_raw = request.POST.getlist('monto')
        tipos = request.POST.getlist('tipo')  # 'debe' | 'haber'

        # ── Manual lines (non-bank accounts) ──────────────────────────────
        lineas = []
        errores = []
        for i, (cid, monto_str, tipo) in enumerate(zip(cuenta_ids, montos_raw, tipos), 1):
            try:
                cid = int(cid)
                monto = Decimal(monto_str.strip().replace('.', '').replace(',', '.'))
                if monto <= 0:
                    continue
                cuenta = PlanCuentas.objects.get(pk=cid, activa=True, acepta_movimientos=True)
                if tipo not in ('debe', 'haber'):
                    tipo = 'debe'
                lineas.append({'cuenta': cuenta, 'monto': monto, 'tipo': tipo})
            except (ValueError, InvalidOperation, PlanCuentas.DoesNotExist):
                errores.append(f'Fila {i}: datos inválidos.')

        # ── Auto-include bank accounts from saldo_inicial ─────────────────
        from apps.tesoreria.models import CuentaBancaria
        for cb in CuentaBancaria.objects.filter(activa=True, cuenta_contable__isnull=False).select_related('cuenta_contable'):
            if cb.saldo_inicial and cb.saldo_inicial != 0:
                lineas.append({
                    'cuenta': cb.cuenta_contable,
                    'monto': abs(cb.saldo_inicial),
                    'tipo': 'debe' if cb.saldo_inicial > 0 else 'haber',
                })

        if errores:
            for e in errores:
                messages.warning(request, e)

        if not lineas:
            messages.error(request, 'No hay líneas con montos válidos.')
            return redirect('contabilidad:apertura')

        # Check confirmed asiento can't be replaced
        existente = self._get_apertura()
        if existente and existente.estado == 'confirmado':
            messages.error(
                request,
                f'Ya existe un asiento de apertura confirmado ({existente.numero}). '
                'Anúlelo desde el Libro Diario antes de reemplazarlo.'
            )
            return redirect('contabilidad:apertura')

        total_debe = sum(l['monto'] for l in lineas if l['tipo'] == 'debe')
        total_haber = sum(l['monto'] for l in lineas if l['tipo'] == 'haber')
        diferencia = total_debe - total_haber  # positive → more assets than liabilities

        # Need patrimonio account if entry doesn't balance
        if diferencia != 0 and (not config or not config.cuenta_patrimonio_apertura):
            messages.error(
                request,
                'El asiento no cuadra y no hay cuenta de Patrimonio / Resultados Acumulados '
                'configurada en Configuración Contable. Configúrela primero o ingrese saldos que cuadren manualmente.'
            )
            return redirect('contabilidad:apertura')

        # Delete existing borrador
        if existente:
            existente.delete()

        # Create new apertura asiento
        apertura = AsientoContable.objects.create(
            fecha=fecha,
            descripcion='Saldos de Apertura',
            tipo='apertura',
            estado='borrador',
        )

        for i, l in enumerate(lineas, start=1):
            LineaAsiento.objects.create(
                asiento=apertura,
                cuenta=l['cuenta'],
                debe=l['monto'] if l['tipo'] == 'debe' else Decimal('0'),
                haber=l['monto'] if l['tipo'] == 'haber' else Decimal('0'),
                descripcion='Saldo de apertura',
                orden=i,
            )

        # Auto-balance line → cuenta_patrimonio_apertura
        if diferencia != 0 and config and config.cuenta_patrimonio_apertura:
            LineaAsiento.objects.create(
                asiento=apertura,
                cuenta=config.cuenta_patrimonio_apertura,
                debe=Decimal('0') if diferencia > 0 else abs(diferencia),
                haber=diferencia if diferencia > 0 else Decimal('0'),
                descripcion='Patrimonio / Resultados Acumulados (cuadre automático)',
                orden=len(lineas) + 1,
            )

        messages.success(
            request,
            f'Asiento de apertura {apertura.numero} guardado en borrador con {len(lineas)} líneas. '
            'Revíselo y confírmelo para que aparezca en los reportes.'
        )
        return redirect('contabilidad:apertura')
