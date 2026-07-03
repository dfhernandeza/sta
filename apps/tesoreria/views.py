from django.views.generic import ListView, CreateView, UpdateView, TemplateView, View, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.utils.dateparse import parse_date
from apps.core.mixins import GestionMixin, AppPermisoMixin

class TesoreriaMixin(AppPermisoMixin):
    app_name = 'tesoreria'

from .models import Banco, CuentaBancaria, MovimientoBancario


def _movimientos_filtrados(parametros):
    qs = MovimientoBancario.objects.select_related(
        'cuenta__banco',
        'cuenta__cuenta_contable',
        'cuenta_contable',
        'proyecto',
        'conciliado_por',
    )
    cuenta_id = parametros.get('cuenta')
    if cuenta_id:
        qs = qs.filter(cuenta_id=cuenta_id)
    tipo = parametros.get('tipo')
    if tipo in {'ingreso', 'egreso'}:
        qs = qs.filter(tipo=tipo)
    conciliado = parametros.get('conciliado')
    if conciliado == '1':
        qs = qs.filter(conciliado=True)
    elif conciliado == '0':
        qs = qs.filter(conciliado=False)
    desde = parse_date(parametros.get('desde', ''))
    if desde:
        qs = qs.filter(fecha__gte=desde)
    hasta = parse_date(parametros.get('hasta', ''))
    if hasta:
        qs = qs.filter(fecha__lte=hasta)
    return qs


class TesoreriaResumenView(TesoreriaMixin, TemplateView):
    template_name = 'admin/tesoreria/resumen.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        cuentas = CuentaBancaria.objects.select_related('banco').filter(activa=True)
        ctx['cuentas'] = cuentas
        ctx['saldo_total'] = sum(c.saldo_actual for c in cuentas)
        ctx['movimientos_recientes'] = MovimientoBancario.objects.select_related(
            'cuenta__banco'
        ).order_by('-fecha', '-creado_en')[:20]
        ctx['titulo'] = 'Tesorería'
        return ctx


class BancoListView(TesoreriaMixin, ListView):
    model = Banco
    template_name = 'admin/tesoreria/banco_list.html'
    context_object_name = 'bancos'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Bancos'
        return ctx


class BancoCreateView(TesoreriaMixin, CreateView):
    model = Banco
    template_name = 'admin/tesoreria/banco_form.html'
    fields = ['nombre', 'codigo']
    success_url = reverse_lazy('tesoreria:banco_list')

    def form_valid(self, form):
        messages.success(self.request, 'Banco registrado.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Nuevo Banco'
        return ctx
    
class BancoUpdateView(TesoreriaMixin, UpdateView):
    model = Banco
    template_name = 'admin/tesoreria/banco_form.html'
    fields = ['nombre', 'codigo']
    success_url = reverse_lazy('tesoreria:banco_list')

    def form_valid(self, form):
        messages.success(self.request, 'Banco actualizado.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f'Editar Banco {self.object.nombre}'
        return ctx


class CuentaBancariaListView(TesoreriaMixin, ListView):
    model = CuentaBancaria
    template_name = 'admin/tesoreria/cuenta_list.html'
    context_object_name = 'cuentas'

    def get_queryset(self):
        return CuentaBancaria.objects.select_related('banco')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Cuentas Bancarias'
        return ctx


class CuentaBancariaCreateView(TesoreriaMixin, CreateView):
    model = CuentaBancaria
    template_name = 'admin/tesoreria/cuenta_form.html'
    fields = ['banco', 'numero', 'tipo', 'descripcion', 'saldo_inicial', 'activa' , 'cuenta_contable']
    success_url = reverse_lazy('tesoreria:cuenta_list')

    def form_valid(self, form):
        messages.success(self.request, 'Cuenta bancaria creada.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Nueva Cuenta Bancaria'
        return ctx


class CuentaBancariaUpdateView(TesoreriaMixin, UpdateView):
    model = CuentaBancaria
    template_name = 'admin/tesoreria/cuenta_form.html'
    fields = ['banco', 'numero', 'tipo', 'descripcion', 'saldo_inicial', 'activa' , 'cuenta_contable']
    success_url = reverse_lazy('tesoreria:cuenta_list')

    def form_valid(self, form):
        messages.success(self.request, 'Cuenta bancaria actualizada.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f'Editar Cuenta {self.object.numero}'
        return ctx


class MovimientoListView(TesoreriaMixin, ListView):
    model = MovimientoBancario
    template_name = 'admin/tesoreria/movimiento_list.html'
    context_object_name = 'movimientos'
    paginate_by = 30

    def get_queryset(self):
        return _movimientos_filtrados(self.request.GET).order_by(
            '-fecha',
            '-creado_en',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['cuentas'] = CuentaBancaria.objects.select_related('banco').filter(activa=True)
        ctx['titulo'] = 'Movimientos Bancarios'
        return ctx


class MovimientosExcelView(TesoreriaMixin, View):
    def get(self, request):
        from io import BytesIO

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        movimientos = list(
            _movimientos_filtrados(request.GET).order_by('fecha', 'creado_en')
        )
        total_ingresos = sum(
            (movimiento.monto for movimiento in movimientos if movimiento.tipo == 'ingreso'),
            0,
        )
        total_egresos = sum(
            (movimiento.monto for movimiento in movimientos if movimiento.tipo == 'egreso'),
            0,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Movimientos'
        headers = [
            'ID', 'Fecha', 'Banco', 'N° Cuenta', 'Tipo de cuenta',
            'Tipo movimiento', 'Monto', 'Descripción', 'Documento',
            'Proyecto', 'Cuenta contable', 'Conciliado',
            'Fecha conciliación', 'Creado el',
        ]
        ultima_columna = get_column_letter(len(headers))
        ws.merge_cells(f'A1:{ultima_columna}1')
        ws['A1'] = 'Movimientos bancarios'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor='1F3864')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 26

        filtros = []
        for clave, etiqueta in [
            ('cuenta', 'Cuenta'),
            ('tipo', 'Tipo'),
            ('conciliado', 'Conciliación'),
            ('desde', 'Desde'),
            ('hasta', 'Hasta'),
        ]:
            valor = request.GET.get(clave)
            if valor:
                filtros.append(f'{etiqueta}: {valor}')
        ws.merge_cells(f'A2:{ultima_columna}2')
        ws['A2'] = 'Filtros: ' + (', '.join(filtros) if filtros else 'Todos los movimientos')
        ws['A2'].font = Font(italic=True, color='666666')

        ws['A3'] = 'Total ingresos'
        ws['B3'] = float(total_ingresos)
        ws['C3'] = 'Total egresos'
        ws['D3'] = float(total_egresos)
        ws['E3'] = 'Saldo neto'
        ws['F3'] = float(total_ingresos - total_egresos)
        for celda in ('A3', 'C3', 'E3'):
            ws[celda].font = Font(bold=True)
        for celda in ('B3', 'D3', 'F3'):
            ws[celda].number_format = '$#,##0.00'
            ws[celda].font = Font(bold=True)

        fila_encabezado = 5
        for columna, encabezado in enumerate(headers, start=1):
            celda = ws.cell(row=fila_encabezado, column=columna, value=encabezado)
            celda.font = Font(bold=True, color='FFFFFF')
            celda.fill = PatternFill('solid', fgColor='4472C4')
            celda.alignment = Alignment(horizontal='center', vertical='center')

        for fila, movimiento in enumerate(movimientos, start=fila_encabezado + 1):
            cuenta_contable = movimiento.cuenta_contable
            creado_el = timezone.localtime(movimiento.creado_en).replace(tzinfo=None)
            valores = [
                movimiento.pk,
                movimiento.fecha,
                movimiento.cuenta.banco.nombre,
                movimiento.cuenta.numero,
                movimiento.cuenta.get_tipo_display(),
                movimiento.get_tipo_display(),
                float(movimiento.monto),
                movimiento.descripcion,
                movimiento.documento,
                str(movimiento.proyecto) if movimiento.proyecto else '',
                (
                    f'{cuenta_contable.codigo} - {cuenta_contable.nombre}'
                    if cuenta_contable else ''
                ),
                'Sí' if movimiento.conciliado else 'No',
                movimiento.fecha_conciliacion,
                creado_el,
            ]
            for columna, valor in enumerate(valores, start=1):
                celda = ws.cell(row=fila, column=columna, value=valor)
                if columna in (2, 13):
                    celda.number_format = 'DD/MM/YYYY'
                elif columna == 14:
                    celda.number_format = 'DD/MM/YYYY HH:MM'
                elif columna == 7:
                    celda.number_format = '$#,##0.00'
                    celda.alignment = Alignment(horizontal='right')

        anchos = [10, 13, 24, 18, 18, 18, 18, 48, 18, 30, 38, 14, 18, 20]
        for columna, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(columna)].width = ancho
        ws.freeze_panes = f'A{fila_encabezado + 1}'
        ws.auto_filter.ref = (
            f'A{fila_encabezado}:{ultima_columna}'
            f'{max(fila_encabezado + len(movimientos), fila_encabezado)}'
        )

        salida = BytesIO()
        wb.save(salida)
        salida.seek(0)
        response = HttpResponse(
            salida.getvalue(),
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
        )
        fecha = timezone.localdate().strftime('%Y%m%d')
        response['Content-Disposition'] = (
            f'attachment; filename="movimientos_bancarios_{fecha}.xlsx"'
        )
        return response


class MovimientoCreateView(TesoreriaMixin, CreateView):
    model = MovimientoBancario
    template_name = 'admin/tesoreria/movimiento_form.html'
    fields = ['cuenta', 'fecha', 'tipo', 'monto', 'descripcion', 'documento', 'cuenta_contable', 'proyecto']
    success_url = reverse_lazy('tesoreria:movimiento_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        movimiento = self.object
        # Generar asiento automáticamente si tiene cuenta bancaria contable y cuenta contable de contrapartida
        if movimiento.cuenta.cuenta_contable and movimiento.cuenta_contable:
            try:
                from apps.contabilidad.utils import generar_asiento_movimiento_bancario
                asiento = generar_asiento_movimiento_bancario(movimiento, usuario=self.request.user)
                if asiento:
                    messages.success(self.request, f'Movimiento registrado. Asiento {asiento.numero} generado en borrador.')
                    return response
            except Exception:
                pass
        messages.success(self.request, 'Movimiento registrado.')
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Nuevo Movimiento Bancario'
        return ctx


class MovimientoUpdateView(TesoreriaMixin, UpdateView):
    model = MovimientoBancario
    template_name = 'admin/tesoreria/movimiento_form.html'
    fields = ['cuenta', 'fecha', 'tipo', 'monto', 'descripcion', 'documento', 'cuenta_contable', 'proyecto']
    success_url = reverse_lazy('tesoreria:movimiento_list')

    def dispatch(self, request, *args, **kwargs):
        movimiento = self.get_object()
        if getattr(movimiento, 'anticipo_proveedor', None):
            messages.error(
                request,
                'No se puede editar este movimiento porque corresponde al pago de un anticipo a proveedor.'
            )
            return redirect('tesoreria:movimiento_list')
        if getattr(movimiento, 'anticipo_laboral', None):
            messages.error(
                request,
                'No se puede editar este movimiento porque corresponde al pago de un anticipo laboral.'
            )
            return redirect('tesoreria:movimiento_list')
        if getattr(movimiento, 'remuneracion_pagada', None):
            messages.error(
                request,
                'No se puede editar este movimiento porque corresponde al pago de una remuneración.'
            )
            return redirect('tesoreria:movimiento_list')
        if getattr(movimiento, 'declaracion_previsional', None):
            messages.error(
                request,
                'No se puede editar este movimiento porque corresponde al pago de una declaración previsional.'
            )
            return redirect('tesoreria:movimiento_list')
        if movimiento.conciliado:
            messages.error(
                request,
                'No se puede editar un movimiento conciliado. Revierta la conciliación antes de modificarlo.'
            )
            return redirect('tesoreria:movimiento_list')
        asientos_confirmados = movimiento.asientos.filter(estado='confirmado')
        if asientos_confirmados.exists():
            numeros = ', '.join(a.numero for a in asientos_confirmados)
            messages.error(
                request,
                f'No se puede editar este movimiento: tiene asiento(s) confirmado(s) asociado(s): {numeros}. '
                'Anule el asiento antes de modificar el movimiento.'
            )
            return redirect('tesoreria:movimiento_list')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, 'Movimiento actualizado.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Editar Movimiento'
        return ctx


class MovimientoDeleteView(TesoreriaMixin, DeleteView):
    model = MovimientoBancario
    template_name = 'admin/confirm_delete.html'
    success_url = reverse_lazy('tesoreria:movimiento_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['cancel_url'] = reverse_lazy('tesoreria:movimiento_list')
        return ctx

    def dispatch(self, request, *args, **kwargs):
        movimiento = self.get_object()
        if movimiento.conciliado:
            messages.error(
                request,
                'No se puede eliminar un movimiento conciliado. Revierta la conciliación antes de eliminarlo.'
            )
            return redirect('tesoreria:movimiento_list')
        if movimiento.cuentas_pagar.exists():
            messages.error(
                request,
                'No se puede eliminar este movimiento directamente porque está vinculado a una Cuenta por Pagar. '
                'Use la acción "Anular" desde Cuentas por Pagar para revertir el pago y limpiar el saldo.'
            )
            return redirect('tesoreria:movimiento_list')
        if getattr(movimiento, 'anticipo_proveedor', None):
            messages.error(
                request,
                'No se puede eliminar este movimiento directamente porque corresponde a un anticipo a proveedor. '
                'Use la acción "Eliminar" desde Anticipos a Proveedores.'
            )
            return redirect('tesoreria:movimiento_list')
        if getattr(movimiento, 'anticipo_laboral', None):
            messages.error(
                request,
                'No se puede eliminar este movimiento directamente porque corresponde a un anticipo laboral. '
                'Use la acción "Eliminar" desde Anticipos Laborales.'
            )
            return redirect('tesoreria:movimiento_list')
        if getattr(movimiento, 'remuneracion_pagada', None):
            messages.error(
                request,
                'No se puede eliminar directamente un pago de remuneración. '
                'Use la eliminación/reversión desde RRHH.'
            )
            return redirect('tesoreria:movimiento_list')
        if getattr(movimiento, 'declaracion_previsional', None):
            messages.error(
                request,
                'No se puede eliminar directamente un pago previsional porque '
                'dejaría la declaración inconsistente.'
            )
            return redirect('tesoreria:movimiento_list')
        asientos_confirmados = movimiento.asientos.filter(estado='confirmado')
        if asientos_confirmados.exists():
            numeros = ', '.join(a.numero for a in asientos_confirmados)
            messages.error(
                request,
                f'No se puede eliminar este movimiento: tiene asiento(s) confirmado(s): {numeros}. '
                'Anule el asiento antes de eliminar el movimiento.'
            )
            return redirect('tesoreria:movimiento_list')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        # Anular asientos en borrador vinculados antes de eliminar el movimiento
        movimiento = self.get_object()
        movimiento.asientos.filter(estado='borrador').update(estado='anulado')
        messages.success(self.request, 'Movimiento eliminado.')
        return super().form_valid(form)


class MovimientoConciliarView(TesoreriaMixin, View):
    def post(self, request, pk):
        movimiento = get_object_or_404(MovimientoBancario, pk=pk)
        if movimiento.conciliado:
            messages.info(request, 'Este movimiento ya está conciliado.')
        else:
            movimiento.conciliado = True
            movimiento.fecha_conciliacion = timezone.localdate()
            movimiento.conciliado_por = request.user
            movimiento.save(update_fields=['conciliado', 'fecha_conciliacion', 'conciliado_por'])
            messages.success(request, 'Movimiento conciliado.')
        return redirect('tesoreria:movimiento_list')


class MovimientoDesconciliarView(TesoreriaMixin, View):
    def post(self, request, pk):
        movimiento = get_object_or_404(MovimientoBancario, pk=pk)
        if not movimiento.conciliado:
            messages.info(request, 'Este movimiento no está conciliado.')
        else:
            movimiento.conciliado = False
            movimiento.fecha_conciliacion = None
            movimiento.conciliado_por = None
            movimiento.save(update_fields=['conciliado', 'fecha_conciliacion', 'conciliado_por'])
            messages.success(request, 'Conciliación revertida.')
        return redirect('tesoreria:movimiento_list')


class GenerarAsientoMovimientoView(TesoreriaMixin, View):
    def post(self, request, pk):
        from apps.contabilidad.utils import generar_asiento_movimiento_bancario, get_config
        movimiento = get_object_or_404(MovimientoBancario, pk=pk)
        asiento_activo = movimiento.asientos.exclude(estado='anulado').first()
        if asiento_activo:
            messages.info(request, f'Este movimiento ya tiene un asiento: {asiento_activo.numero}.')
            return redirect('contabilidad:asiento_detail', pk=asiento_activo.pk)
        if not get_config():
            messages.warning(request, 'Configure primero las cuentas contables antes de generar asientos.')
            return redirect('contabilidad:configuracion')
        if not movimiento.cuenta.cuenta_contable:
            messages.warning(request, 'La cuenta bancaria no tiene una cuenta contable asignada. Edite la cuenta bancaria primero.')
            return redirect('tesoreria:movimiento_list')
        if not movimiento.cuenta_contable:
            messages.warning(request, 'El movimiento no tiene cuenta contable de contrapartida asignada.')
            return redirect('tesoreria:movimiento_list')
        asiento = generar_asiento_movimiento_bancario(movimiento, usuario=request.user)
        if asiento:
            messages.success(request, f'Asiento {asiento.numero} generado en borrador.')
            return redirect('contabilidad:asiento_detail', pk=asiento.pk)
        messages.error(request, 'No se pudo generar el asiento.')
        return redirect('tesoreria:movimiento_list')
