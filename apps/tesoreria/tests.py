from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.test import Client, TestCase, override_settings
from django.urls import reverse

from apps.accounts.models import CustomUser
from .models import Banco, CuentaBancaria, MovimientoBancario


@override_settings(SECURE_SSL_REDIRECT=False)
class MovimientosExcelViewTest(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(
            'tesoreria_excel',
            password='pass',
            app_permisos=['tesoreria'],
        )
        self.client_http = Client()
        self.client_http.force_login(self.user)
        banco = Banco.objects.create(nombre='Banco Excel', codigo='BEX')
        self.cuenta = CuentaBancaria.objects.create(
            banco=banco,
            numero='123456',
            tipo='corriente',
            saldo_inicial=Decimal('1000000'),
        )
        self.ingreso = MovimientoBancario.objects.create(
            cuenta=self.cuenta,
            fecha=date(2026, 7, 1),
            tipo='ingreso',
            monto=Decimal('250000'),
            descripcion='Ingreso de prueba',
            documento='ING-001',
        )
        self.egreso = MovimientoBancario.objects.create(
            cuenta=self.cuenta,
            fecha=date(2026, 7, 2),
            tipo='egreso',
            monto=Decimal('75000'),
            descripcion='Egreso de prueba',
            documento='EGR-001',
            conciliado=True,
            fecha_conciliacion=date(2026, 7, 3),
        )
        self.url = reverse('tesoreria:movimiento_excel')

    def _libro(self, response):
        return openpyxl.load_workbook(BytesIO(response.content))

    def test_descarga_excel_con_todos_los_movimientos(self):
        response = self.client_http.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn(
            'attachment; filename="movimientos_bancarios_',
            response['Content-Disposition'],
        )
        ws = self._libro(response)['Movimientos']
        self.assertEqual(ws['A1'].value, 'Movimientos bancarios')
        self.assertEqual(ws['A5'].value, 'ID')
        ids = {ws['A6'].value, ws['A7'].value}
        self.assertEqual(ids, {self.ingreso.pk, self.egreso.pk})
        self.assertEqual(ws['B3'].value, 250000)
        self.assertEqual(ws['D3'].value, 75000)
        self.assertEqual(ws['F3'].value, 175000)

    def test_excel_respeta_filtros_del_listado(self):
        response = self.client_http.get(
            self.url,
            {'tipo': 'ingreso', 'desde': '2026-07-01', 'hasta': '2026-07-01'},
        )

        ws = self._libro(response)['Movimientos']
        self.assertEqual(ws['A6'].value, self.ingreso.pk)
        self.assertIsNone(ws['A7'].value)
        self.assertEqual(ws['B3'].value, 250000)
        self.assertEqual(ws['D3'].value, 0)

    def test_listado_muestra_boton_excel_con_filtros(self):
        response = self.client_http.get(
            reverse('tesoreria:movimiento_list'),
            {'tipo': 'egreso', 'conciliado': '1'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Descargar Excel')
        self.assertContains(response, 'tipo=egreso')
        self.assertContains(response, 'conciliado=1')
